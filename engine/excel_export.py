"""
excel_export.py — Export timetables to a colour-coded Excel workbook.

Colour legend:
    Green  (#E8F5E9) — Lecture  [L]
    Blue   (#E3F2FD) — Tutorial [T]
    Orange (#FFF3E0) — Practical / Co-Fac [P]
    Grey   (#F5F5F5) — Free slot

Public API:
    create_timetables_excel(section_timetables, faculty_timetables) -> bytes
"""

import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
SLOTS = [
    "S1 (9:00-9:55)", "S2 (9:55-10:50)", "S3 (11:05-12:00)",
    "S4 (12:00-12:50)", "S5 (1:45-2:40)", "S6 (2:40-3:35)", "S7 (3:35-4:30)",
]

# Fills
_SEC_HEADER_FILL  = PatternFill("solid", fgColor="37474F")   # dark blue-grey
_FAC_HEADER_FILL  = PatternFill("solid", fgColor="1565C0")   # dark blue
_DAY_FILL         = PatternFill("solid", fgColor="546E7A")   # medium blue-grey
_LECTURE_FILL     = PatternFill("solid", fgColor="E8F5E9")   # light green
_TUTORIAL_FILL    = PatternFill("solid", fgColor="E3F2FD")   # light blue
_PRACTICAL_FILL   = PatternFill("solid", fgColor="FFF3E0")   # light orange
_FREE_FILL        = PatternFill("solid", fgColor="F5F5F5")   # light grey
_ALT_FREE_FILL    = PatternFill("solid", fgColor="FAFAFA")   # off-white

# Fonts
_HEADER_FONT      = Font(bold=True, color="FFFFFF", size=9)
_DAY_FONT         = Font(bold=True, color="FFFFFF", size=9)
_TITLE_FONT       = Font(bold=True, size=11)
_CELL_FONT        = Font(size=8)

# Borders
_thin = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# Alignment
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Column widths (col A = Day, cols B-H = slots)
_COL_A_WIDTH = 14
_COL_SLOT_WIDTH = 26

# Row heights
_TITLE_ROW_H  = 22
_HEADER_ROW_H = 28
_DATA_ROW_H   = 58


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _cell_fill(raw: str, row_idx: int) -> PatternFill:
    """Return the appropriate fill based on the event type tag in the cell text."""
    if not raw:
        return _ALT_FREE_FILL if row_idx % 2 == 0 else _FREE_FILL
    if "[L]" in raw:
        return _LECTURE_FILL
    if "[T]" in raw:
        return _TUTORIAL_FILL
    if "[P]" in raw or "Co-Fac" in raw:
        return _PRACTICAL_FILL
    return _ALT_FREE_FILL if row_idx % 2 == 0 else _FREE_FILL


def _write_grid(ws, title: str, grid: list, header_fill: PatternFill):
    """Write a 5-day x 7-slot timetable grid into worksheet *ws*.

    Layout:
        Row 1  -- merged title (A1:H1)
        Row 2  -- header: Day | S1 | S2 | ... | S7
        Rows 3-7 -- data rows (one per day)
    """
    # --- Title row ---
    ws.merge_cells("A1:H1")
    tc = ws["A1"]
    tc.value = title
    tc.font = _TITLE_FONT
    tc.alignment = _CENTER
    tc.fill = header_fill
    ws.row_dimensions[1].height = _TITLE_ROW_H

    # --- Header row (row 2) ---
    day_hdr = ws.cell(row=2, column=1, value="Day")
    day_hdr.fill = header_fill
    day_hdr.font = _HEADER_FONT
    day_hdr.alignment = _CENTER
    day_hdr.border = _BORDER

    for col_i, slot in enumerate(SLOTS, start=2):
        c = ws.cell(row=2, column=col_i, value=slot)
        c.fill = header_fill
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.border = _BORDER
    ws.row_dimensions[2].height = _HEADER_ROW_H

    # --- Data rows ---
    for d_idx, day in enumerate(DAYS):
        row_num = d_idx + 3

        # Day label
        dc = ws.cell(row=row_num, column=1, value=day)
        dc.fill = _DAY_FILL
        dc.font = _DAY_FONT
        dc.alignment = _CENTER
        dc.border = _BORDER

        for t_idx in range(7):
            raw = ""
            if d_idx < len(grid) and t_idx < len(grid[d_idx]):
                raw = grid[d_idx][t_idx] or ""
            cell = ws.cell(row=row_num, column=t_idx + 2, value=raw)
            cell.fill = _cell_fill(raw, d_idx)
            cell.font = _CELL_FONT
            cell.alignment = _CENTER
            cell.border = _BORDER

        ws.row_dimensions[row_num].height = _DATA_ROW_H

    # --- Column widths ---
    ws.column_dimensions["A"].width = _COL_A_WIDTH
    for col_i in range(2, 9):
        ws.column_dimensions[get_column_letter(col_i)].width = _COL_SLOT_WIDTH

    # Freeze the header rows so they stay visible when scrolling
    ws.freeze_panes = "B3"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def create_timetables_excel(
    section_timetables: dict,
    faculty_timetables: dict | None = None,
    academic_year: str = "",
) -> bytes:
    """Generate a colour-coded Excel workbook and return it as bytes.

    Parameters
    ----------
    section_timetables : dict
        {section_name: 5x7 grid of cell strings}
    faculty_timetables : dict, optional
        {faculty_name: 5x7 grid of cell strings}
    academic_year : str, optional
        Appended to sheet titles, e.g. "2025-26".

    Returns
    -------
    bytes
        Raw .xlsx content suitable for st.download_button.
    """
    wb = Workbook()
    # Remove the default blank sheet
    wb.remove(wb.active)

    year_suffix = f" ({academic_year})" if academic_year else ""

    # --- Section sheets ---
    for sec in sorted(section_timetables.keys()):
        sheet_name = f"Sec {sec}"[:31]   # Excel sheet name max 31 chars
        ws = wb.create_sheet(title=sheet_name)
        _write_grid(
            ws,
            title=f"Section Timetable - {sec}{year_suffix}",
            grid=section_timetables[sec],
            header_fill=_SEC_HEADER_FILL,
        )

    # --- Faculty sheets ---
    if faculty_timetables:
        for fac in sorted(faculty_timetables.keys()):
            sheet_name = f"Fac {fac}"[:31]
            ws = wb.create_sheet(title=sheet_name)
            _write_grid(
                ws,
                title=f"Faculty Timetable - {fac}{year_suffix}",
                grid=faculty_timetables[fac],
                header_fill=_FAC_HEADER_FILL,
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
