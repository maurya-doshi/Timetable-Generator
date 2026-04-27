import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
from db import get_db

st.set_page_config(page_title="Input Data", page_icon="📋", layout="wide")

st.title("📋 Input Data — Faculty & Subject Allocation")

st.markdown(
    """
    Upload an **Excel file** (`.xlsx`) containing **two sheets**:

    1. **Faculty_Assignments** – faculty allotment (two‑row header format)
    2. **Courses** – course details with L/T/P, lab flags, and elective info

    ### Faculty_Assignments sheet format:
    | Sr No. | Name | Designation | Subject |       |       |       |       |       | Lab   |       |       |       | ... |
    |--------|------|-------------|---------|-------|-------|-------|-------|-------|-------|-------|-------|-------|-----|
    |        |      |             | S1      | Sem S1| S2    | Sem S2| S3    | Sem S3| L1    | Sem L1| L2    | Sem L2| ... |
    | 1      | ...  | ...         | CS301   | 3     | CS302 | 3A    | CS302 | 3B    | ...   | ...   | ...   | ...   | ... |

    *(**Note**: In the `Sem` columns, you can specify exact sections like `3A` or `3A, 3B` to assign faculty to specific sections. If you just put the semester number like `3`, the system will automatically distribute teachers round-robin across all sections.)*

    ### Courses sheet format:
    | Course Code | Course Name | L | T | P | Lecture in Lab? | Tutorial in Lab? | Semester | Elective |
    |-------------|-------------|---|---|---|-----------------|------------------|----------|----------|
    | 24CS32      | Digital Design and Computer Organization | 3 | 0 | 1 | No | No | 3 | No |
    """
)

st.divider()

# ---------------------------------------------------------------------------
# Semester selection (only for faculty data)
# ---------------------------------------------------------------------------
semester = st.radio("Which semester is this file for?", ["Odd", "Even"], horizontal=True)

# ---------------------------------------------------------------------------
# Excel parser – Faculty_Assignments sheet
# ---------------------------------------------------------------------------
def find_header_row(ws):
    """Find row containing 'Sr No.', 'Name', 'Designation'."""
    for row_idx in range(1, min(ws.max_row + 1, 100)):
        row_values = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[row_idx]]
        has_sr = any(("sr" in v and "no" in v) for v in row_values)
        has_name = any("name" in v for v in row_values)
        has_design = any("designation" in v or "design" in v for v in row_values)
        if has_sr and has_name and has_design:
            return row_idx
    return None

def parse_faculty_sheet(ws):
    """Parse the Faculty_Assignments sheet (two‑row header)."""
    debug = {}
    header_row = find_header_row(ws)
    if header_row is None:
        return None, debug, "Faculty_Assignments: Could not find header row."

    sub_header_row = header_row + 1
    sub_header = [cell.value for cell in ws[sub_header_row]]

    subject_cols = []      # list of (subject_col_index, semester_col_index)
    lab_cols = []          # list of (lab_col_index, semester_col_index)

    col_idx = 3  # skip first three columns (Sr No., Name, Designation)
    while col_idx < len(sub_header):
        cell_val = str(sub_header[col_idx]).strip() if sub_header[col_idx] else ""
        if cell_val.startswith("S") and not cell_val.startswith("Sem"):
            if col_idx + 1 < len(sub_header):
                semester_val = str(sub_header[col_idx + 1]).strip() if sub_header[col_idx + 1] else ""
                if "Sem" in semester_val:
                    subject_cols.append((col_idx, col_idx + 1))
                    col_idx += 2
                    continue
        elif cell_val.startswith("L"):
            if col_idx + 1 < len(sub_header):
                semester_val = str(sub_header[col_idx + 1]).strip() if sub_header[col_idx + 1] else ""
                if "Sem" in semester_val:
                    lab_cols.append((col_idx, col_idx + 1))
                    col_idx += 2
                    continue
        col_idx += 1

    data_start = sub_header_row + 1
    records = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not any(cell for cell in row):
            continue
        sl_no_raw = row[0] if len(row) > 0 else None
        name_raw = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        designation = str(row[2]).strip() if len(row) > 2 and row[2] else ""

        if not name_raw or name_raw.lower() in ("none", "", "name of faculty"):
            continue

        subjects = []
        for subj_col, sem_col in subject_cols:
            if subj_col < len(row) and row[subj_col]:
                subj_code = str(row[subj_col]).strip()
                sem_val = str(row[sem_col]).strip() if sem_col < len(row) and row[sem_col] else ""
                if subj_code and subj_code.lower() != "none" and subj_code != name_raw:
                    subjects.append({"code": subj_code, "semester": sem_val})

        labs = []
        for lab_col, sem_col in lab_cols:
            if lab_col < len(row) and row[lab_col]:
                lab_code = str(row[lab_col]).strip()
                sem_val = str(row[sem_col]).strip() if sem_col < len(row) and row[sem_col] else ""
                if lab_code and lab_code.lower() != "none" and lab_code != name_raw:
                    labs.append({"code": lab_code, "semester": sem_val})

        records.append({
            "sl_no": sl_no_raw,
            "name": name_raw,
            "designation": designation,
            "subjects": subjects,
            "labs": labs,
        })
    return records, debug, None

def parse_courses_sheet(ws):
    """Parse the Courses sheet – robust header matching."""
    # Read the first row (row 1) as headers
    headers = []
    for cell in ws[1]:
        val = str(cell.value).strip() if cell.value else ""
        headers.append(val)
    
    # Expected column names (can be partial matches)
    expected_patterns = {
        "Course Code": ["course code", "course", "code"],
        "Course Name": ["course name", "name", "title"],
        "L": ["l", "lecture"],
        "T": ["t", "tutorial"],
        "P": ["p", "practical", "lab"],
        "Lecture in Lab?": ["lecture in lab", "lecture lab", "lec in lab"],
        "Tutorial in Lab?": ["tutorial in lab", "tut in lab", "tutorial lab"],
        "Semester": ["semester", "sem"],
        "Elective": ["elective"],
        "AEC": ["ability enhancement", "abiliity enhancement", "aec"],
        "UG_PG": ["ug/pg", "ug", "pg"]
    }
    
    # Map columns by matching patterns
    col_map = {}
    for col_idx, header in enumerate(headers):
        header_lower = header.lower()
        for expected, patterns in expected_patterns.items():
            if expected not in col_map and any(pattern in header_lower for pattern in patterns):
                col_map[expected] = col_idx
                break
    
    # Check for missing required columns
    required = ["Course Code", "Course Name", "L", "T", "P", "Semester"]
    missing = [r for r in required if r not in col_map]
    if missing:
        return None, f"Missing columns: {missing}. Found headers: {headers}"
    
    # Optional columns (Lecture in Lab?, Tutorial in Lab?) – default to "No"
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        
        def get_val(key, default=""):
            idx = col_map.get(key)
            if idx is not None and idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return default
        
        def get_int(key, default=0):
            val = get_val(key, "")
            try:
                return int(float(val)) if val else default
            except:
                return default
        
        course_code = get_val("Course Code")
        if not course_code:
            continue
        
        record = {
            "course_code": course_code,
            "course_name": get_val("Course Name"),
            "L": get_int("L"),
            "T": get_int("T"),
            "P": get_int("P"),
            "lecture_in_lab": get_val("Lecture in Lab?", "No"),
            "tutorial_in_lab": get_val("Tutorial in Lab?", "No"),
            "semester": get_val("Semester"),
            "elective": get_val("Elective", "No"),
            "aec": get_val("AEC", "No"),
            "ug_pg": get_val("UG_PG", "UG"),
        }
        records.append(record)
    
    return records, None

# ---------------------------------------------------------------------------
# File upload
# ---------------------------------------------------------------------------
uploaded_file = st.file_uploader(
    "Upload Excel File (with both sheets)",
    type=["xlsx"],
    help="Upload a .xlsx file containing 'Faculty_Assignments' and 'Courses' sheets.",
)

if uploaded_file is not None:
    try:
        wb = load_workbook(BytesIO(uploaded_file.read()), data_only=True)
    except Exception as e:
        st.error(f"Failed to read Excel file: {e}")
        st.stop()

    # --- Parse Faculty_Assignments sheet ---
    if "Faculty_Assignments" not in wb.sheetnames:
        st.error("Excel file must contain a sheet named 'Faculty_Assignments'.")
        st.stop()
    faculty_ws = wb["Faculty_Assignments"]
    faculty_records, debug, faculty_error = parse_faculty_sheet(faculty_ws)
    if faculty_error:
        st.error(f"Faculty_Assignments error: {faculty_error}")
        with st.expander("Debug info"):
            st.json(debug)
        st.stop()
    st.success(f"✅ Parsed **{len(faculty_records)} faculty record(s)** from Faculty_Assignments.")

    # --- Parse Courses sheet ---
    if "Courses" not in wb.sheetnames:
        st.error("Excel file must contain a sheet named 'Courses'.")
        st.stop()
    courses_ws = wb["Courses"]
    courses_records, courses_error = parse_courses_sheet(courses_ws)
    if courses_error:
        st.error(f"Courses error: {courses_error}")
        st.stop()
    st.success(f"✅ Parsed **{len(courses_records)} course(s)** from Courses.")

    # -----------------------------------------------------------------------
    # Display Faculty Assignments separately
    # -----------------------------------------------------------------------
    st.header("📚 Faculty Assignments")
    st.subheader("Preview")
    if faculty_records:
        header_cols = st.columns([0.5, 2, 1.5, 3, 2])
        header_cols[0].markdown("**Sl.**")
        header_cols[1].markdown("**Faculty Name**")
        header_cols[2].markdown("**Designation**")
        header_cols[3].markdown("**Subjects (Semester)**")
        header_cols[4].markdown("**Labs (Semester)**")
        for rec in faculty_records:
            cols = st.columns([0.5, 2, 1.5, 3, 2])
            cols[0].write(rec["sl_no"] if rec["sl_no"] else "—")
            cols[1].write(rec["name"])
            cols[2].write(rec["designation"])
            subj_str = ", ".join(f"{s['code']} ({s['semester']})" for s in rec["subjects"]) if rec["subjects"] else "—"
            lab_str = ", ".join(f"{l['code']} ({l['semester']})" for l in rec["labs"]) if rec["labs"] else "—"
            cols[3].write(subj_str)
            cols[4].write(lab_str)
    else:
        st.info("No faculty records found.")

    # --- Faculty: Save and Delete buttons (side by side) ---
    col_fac_save, col_fac_delete = st.columns(2)
    with col_fac_save:
        if st.button("💾 Save Faculty to Database", type="primary", key="save_fac"):
            db = get_db()
            collection_name = f"faculty_{semester.lower()}"
            col = db[collection_name]
            docs = []
            for rec in faculty_records:
                docs.append({
                    "sl_no": rec["sl_no"],
                    "name": rec["name"],
                    "designation": rec["designation"],
                    "subjects": rec["subjects"],
                    "labs": rec["labs"],
                    "semester": semester.lower(),
                })
            col.delete_many({})
            if docs:
                col.insert_many(docs)
            st.success(f"✅ Saved {len(docs)} faculty records to `{collection_name}`.")
            st.rerun()
    with col_fac_delete:
        if st.button("🗑️ Delete All Faculty Records", type="secondary", key="del_fac"):
            st.session_state["confirm_fac_delete"] = True
        if st.session_state.get("confirm_fac_delete", False):
            st.warning(f"⚠️ Delete ALL faculty records for **{semester}** semester? This cannot be undone.")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("Yes, delete faculty", key="confirm_fac_yes"):
                    db = get_db()
                    collection_name = f"faculty_{semester.lower()}"
                    result = db[collection_name].delete_many({})
                    st.success(f"🗑️ Deleted {result.deleted_count} faculty records.")
                    st.session_state["confirm_fac_delete"] = False
                    st.rerun()
            with col2:
                if st.button("Cancel", key="confirm_fac_no"):
                    st.session_state["confirm_fac_delete"] = False
                    st.rerun()

    st.divider()

    # -----------------------------------------------------------------------
    # Display Courses separately
    # -----------------------------------------------------------------------
    st.header("📖 Courses")
    st.subheader("Preview")
    if courses_records:
        st.dataframe(courses_records, use_container_width=True)
    else:
        st.info("No course records found.")

    # --- Courses: Save and Delete buttons ---
    col_course_save, col_course_delete = st.columns(2)
    with col_course_save:
        if st.button("💾 Save Courses to Database", type="primary", key="save_courses"):
            db = get_db()
            collection_name = "courses"
            col = db[collection_name]
            col.delete_many({})  # replace all courses
            if courses_records:
                col.insert_many(courses_records)
            st.success(f"✅ Saved {len(courses_records)} courses to `{collection_name}`.")
            st.rerun()
    with col_course_delete:
        if st.button("🗑️ Delete All Course Records", type="secondary", key="del_courses"):
            st.session_state["confirm_course_delete"] = True
        if st.session_state.get("confirm_course_delete", False):
            st.warning("⚠️ Delete ALL course records? This cannot be undone.")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("Yes, delete courses", key="confirm_course_yes"):
                    db = get_db()
                    collection_name = "courses"
                    result = db[collection_name].delete_many({})
                    st.success(f"🗑️ Deleted {result.deleted_count} course records.")
                    st.session_state["confirm_course_delete"] = False
                    st.rerun()
            with col2:
                if st.button("Cancel", key="confirm_course_no"):
                    st.session_state["confirm_course_delete"] = False
                    st.rerun()

    st.divider()

    # -----------------------------------------------------------------------
    # Display current database state (optional)
    # -----------------------------------------------------------------------
    st.subheader("Current Database State")
    db = get_db()
    fac_col = db[f"faculty_{semester.lower()}"]
    fac_existing = list(fac_col.find({}, {"_id": 0}))
    st.write(f"**Faculty ({semester} semester):** {len(fac_existing)} records")
    courses_existing = list(db["courses"].find({}, {"_id": 0}))
    st.write(f"**Courses:** {len(courses_existing)} records")
    if st.checkbox("Show current courses"):
        st.dataframe(courses_existing)

else:
    st.info("📂 Please upload an Excel file to begin.")